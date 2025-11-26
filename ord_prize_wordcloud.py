from openpyxl import load_workbook
import os
import matplotlib as mpl
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
# # from PIL import Image
import numpy as np
from wordcloud import WordCloud, STOPWORDS
from collections import Counter
from matplotlib.colors import LinearSegmentedColormap

def get_text_color(color):
    # color is rgba tuple
    r, g, b, a = color
    luminance = 0.299 * r + 0.587 * g + 0.114 * b
    return 'white' if luminance < 0.5 else 'black'

font_path = 'Geneva.ttf'
font_prop = fm.FontProperties(fname=font_path)

input_dir = '/Users/clange/SynologyDrive/SoundingBoardResearchers/ORD_Prize_2025/ORDPrize_2025_Entries'

row_dict = {
    'gender': 'B2',
    'institution': 'B8',
    'research_discipline': 'B10',
    'position': 'B11',
    'title': 'B16',
    'description': 'B16',
    'aspect_description': 'B25',
    'summary': 'B26',
}

# Define custom colors (Pantone Red 032 U and Pantone 7546 U)
pantone_red = np.array([234, 81, 90]) / 255.0  # RGB normalized to 0-1
pantone_grey = np.array([90, 97, 107]) / 255.0  # RGB normalized to 0-1

# Create custom colormap blending from grey to red
colors_list = [pantone_grey, pantone_red]
n_bins = 20
cmap = LinearSegmentedColormap.from_list('pantone_custom', colors_list, N=n_bins)

worksheets = []
for file in os.listdir(input_dir):
    if file.endswith('.xlsx') and not file.startswith('~$'):
        full_path = os.path.join(input_dir, file)
        wb = load_workbook(full_path)
        ws = wb.active
        worksheets.append(ws)


# Collect gender values and create pie chart
genders = [ws[row_dict['gender']].value for ws in worksheets]
gender_counts = Counter(genders)

# colors = plt.get_cmap('berlin')(np.linspace(0, 1, len(gender_counts)))
colors = cmap(np.linspace(0, 1, len(gender_counts)))
fig, ax = plt.subplots(figsize=(8, 6))
fig.patch.set_alpha(0)
ax.patch.set_alpha(0)
pie = plt.pie(gender_counts.values(), labels=list(gender_counts.keys()), autopct='%1.1f%%', colors=colors, textprops={'fontproperties': font_prop, 'fontsize': 14})
patches, texts, autotexts = pie
for i, autotext in enumerate(autotexts):
    color = patches[i].get_facecolor()
    autotext.set_color(get_text_color(color))
for text in texts:
    text.set_bbox(dict(boxstyle='round,pad=0.3', facecolor='white', alpha=0.7, edgecolor='none'))
plt.title('Gender Distribution', fontproperties=font_prop, fontsize=16)
plt.savefig('gender_pie_chart.pdf', transparent=True)


# Research disciplines and domains
domains_dict = {
    '1': 'Human and Social Sciences',
    '2': 'Mathematics, Natural-\nand Engineering Sciences',
    '3': 'Biology and Medicine',
    '4': 'Interdisciplinary\nResearch',
}

research_disciplines = [ws[row_dict['research_discipline']].value for ws in worksheets]
print(research_disciplines)
# The disciplines have the form '2.6 Environmental Sciences'
# To extract the main domain, we can split by '.' and take the first part
research_domains = [discipline.split('.')[0] if discipline and '.' in discipline else 'Unknown' for discipline in research_disciplines]
domain_counts = Counter(research_domains)
print(domain_counts)
plt.figure(figsize=(8, 6))
labels = [domains_dict.get(key, key) for key in domain_counts.keys()]
# colors = plt.get_cmap('berlin')(np.linspace(0, 1, len(domain_counts)))
colors = cmap(np.linspace(0, 1, len(domain_counts)))
fig, ax = plt.subplots(figsize=(8, 6))
fig.patch.set_alpha(0)
ax.patch.set_alpha(0)
pie = plt.pie(domain_counts.values(), labels=labels, autopct='%1.1f%%', colors=colors, textprops={'fontproperties': font_prop, 'fontsize': 14})
patches, texts, autotexts = pie
for i, autotext in enumerate(autotexts):
    color = patches[i].get_facecolor()
    autotext.set_color(get_text_color(color))
for i, text in enumerate(texts):
    text.set_bbox(dict(boxstyle='round,pad=0.3', facecolor='white', alpha=0.7, edgecolor='none'))
    # Move Mathematics label slightly right to avoid overlap
    if labels[i] == 'Mathematics, Natural-\nand Engineering Sciences':
        x, y = text.get_position()
        text.set_position((x * 1.07, y))
plt.title('Research Domains', fontproperties=font_prop, fontsize=16)
plt.savefig('research_domains_pie_chart.pdf', transparent=True)

# Plot position distribution
positions = [ws[row_dict['position']].value for ws in worksheets]
position_counts = Counter(positions)
print(position_counts)
fig, ax = plt.subplots(figsize=(10, 6))
fig.patch.set_alpha(0)
ax.patch.set_alpha(0)
labels = list(position_counts.keys())
# colors = plt.get_cmap('berlin')(np.linspace(0, 1, len(position_counts)))
colors = cmap(np.linspace(0, 1, len(position_counts)))
pie = plt.pie(position_counts.values(), labels=labels, autopct='%1.1f%%', colors=colors, textprops={'fontproperties': font_prop, 'fontsize': 14})
patches, texts, autotexts = pie
for i, autotext in enumerate(autotexts):
    color = patches[i].get_facecolor()
    autotext.set_color(get_text_color(color))
for text in texts:
    text.set_bbox(dict(boxstyle='round,pad=0.3', facecolor='white', alpha=0.7, edgecolor='none'))
plt.title('Position of Applicants', fontproperties=font_prop, fontsize=16)
plt.savefig('position_pie_chart.pdf', transparent=True)


# Plot institution distribution
institutions = [ws[row_dict['institution']].value for ws in worksheets]
print(institutions)
institution_counts = Counter(institutions)
print(institution_counts)


# Word cloud for descriptions
descriptions = [ws[row_dict['description']].value for ws in worksheets if ws[row_dict['description']].value]
all_text = ' '.join(descriptions).lower()


stopwords= set(STOPWORDS)
stopwords.add("Ugo")
stopwords.add("Foscolo")
stopwords.add("Masonry")
stopwords.add("Obelis")
stopwords.add("reinhardtii")
stopwords.add("marmot")
stopwords.add("proteomics")
stopwords.add("n2pc")
stopwords.add("beast")
stopwords.add("wedowind")
stopwords.add("hcm")
stopwords.add("covid")
stopwords.add("papier")
stopwords.add("wed")
stopwords.add("taming")
stopwords.add("sans")
stopwords.add("chlamydomonas")


#generate the word cloud with parameters
wc = WordCloud(background_color=None,
               mode='RGBA',
               font_path="Geneva.ttf",
               max_words=90,
               width=800,
               height=600,
               min_font_size =20,
               max_font_size=100,
               relative_scaling = 0.5,
               scale=2,
               stopwords=stopwords,
               random_state=1,
               collocations=True,
               colormap=cmap,
            #    colormap="Blues",
            #    colormap="plasma",
            #    colormap="winter",
               normalize_plurals= True)
wc.generate(all_text)
fig, ax = plt.subplots(figsize=(15, 15))
fig.patch.set_alpha(0)
ax.patch.set_alpha(0)
# plt.figure()
plt.imshow(wc, interpolation="bilinear")
plt.axis("off")

#Show the wordcloud
plt.savefig('wordcloud.pdf', transparent=True)
